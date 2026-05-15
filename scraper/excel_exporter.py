import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime


class ExcelExporter:
    def __init__(self, output_path: str):
        self.output_path = output_path
        self.wb = Workbook()
        # Remove default sheet
        self.wb.remove(self.wb.active)
        
    def _style_header(self, ws):
        """Apply styling to header row."""
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths."""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
            
    def add_domain_sheet(self, domain_name: str, leads: list):
        """Add a sheet for a specific domain."""
        ws = self.wb.create_sheet(title=domain_name[:31])  # Excel sheet name max 31 chars
        
        if not leads:
            ws['A1'] = "No data found"
            return
            
        # Convert leads to DataFrame
        df = pd.DataFrame([lead.__dict__ for lead in leads])
        
        # Add metadata columns
        df['extraction_date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        df['extraction_tool'] = 'Local Maps Scraper v1.0'
        
        # Columns to drop entirely from the Excel output
        excluded_cols = ['services', 'hours', 'latitude', 'longitude', 'place_id']
        df = df.drop(columns=[c for c in excluded_cols if c in df.columns])

        # Reorder columns for readability
        priority_cols = ['business_name', 'type', 'domain', 'phone', 'email',
                        'website', 'address', 'city', 'zone', 'rating',
                        'review_count', 'revenue_estimate', 'team_size',
                        'creation_date', 'extraction_date']

        existing_cols = [c for c in priority_cols if c in df.columns]
        other_cols = [c for c in df.columns if c not in priority_cols]
        df = df[existing_cols + other_cols]
        
        # Write to worksheet
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
                
        self._style_header(ws)
        self._auto_adjust_columns(ws)
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Add filter
        ws.auto_filter.ref = ws.dimensions
        
    def add_summary_sheet(self, stats: dict):
        """Add a summary/overview sheet."""
        ws = self.wb.create_sheet(title='Summary', index=0)
        
        ws['A1'] = 'LEAD EXTRACTION REPORT'
        ws['A1'].font = Font(size=16, bold=True, color='366092')
        
        ws['A3'] = 'Generated:'
        ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        ws['A4'] = 'City:'
        ws['B4'] = stats.get('city', 'N/A')
        
        ws['A5'] = 'Total Domains:'
        ws['B5'] = stats.get('total_domains', 0)
        
        ws['A6'] = 'Total Leads Extracted:'
        ws['B6'] = stats.get('total_leads', 0)
        
        ws['A8'] = 'Breakdown by Domain:'
        ws['A8'].font = Font(bold=True)
        
        row = 9
        for domain, count in stats.get('domain_counts', {}).items():
            ws[f'A{row}'] = domain
            ws[f'B{row}'] = count
            row += 1
            
        self._auto_adjust_columns(ws)
        
    def save(self):
        """Save the workbook."""
        self.wb.save(self.output_path)
        print(f"✓ Excel file saved: {self.output_path}")